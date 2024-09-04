from openpyxl import load_workbook
from pandas import read_csv

# csv转化excel
# 打开csv文件
f = open('./21计应301成绩.csv')
data = read_csv(f)#读取
data.to_excel('./21计应301成绩.xlsx')#生成excel

students = load_workbook('./21计应301成绩.xlsx')
active_st = students.active  # 获取活动工作表
print(active_st['K5'].value)# 输出姓名


for row in active_st.iter_rows(min_row=5,max_row=5,min_col=13,max_col=13):
    row[0].value="我真优秀!"
    print(row[0].value)

# num = int(1)
# for col in active_st['K']:
#     print(col.value)
#     if col.value == '欧阳小鹏':
#         chengji = 'M' + str(num)
#         active_st[chengji].value = '我真优秀!!'
#         print(active_st[chengji].value)
#     num += 1


# # 修改单元格值
# active_st['M5'] = '你真棒!!!'
# # 输出修改单元格后的值
# print(active_st['M5'].value)


# 保存文件
students.save('./21计应301成绩.xlsx')