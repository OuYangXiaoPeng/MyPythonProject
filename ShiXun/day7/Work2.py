from openpyxl import load_workbook, Workbook


h = []
h.append('名字')
wb2 = Workbook()
ws2 = wb2.active
ws2.append(h)

wb = load_workbook('./11月考勤统计.xlsx')
ws=wb.active
wwb= load_workbook('./迟到次数月度统计（11月更新）.xlsx')
wws = wwb.active

for row in ws.iter_rows(min_row=2, values_only=True):
    for row2 in wws.iter_rows(min_row=3, values_only=True):
        name1 = row[1]
        times1 = row[-1]
        name2 = row2[1]
        times2 = row2[13]
        if name1 == name2:
            if times1 != times2:
                print('名字为{}次数不匹配,统计表为:{},迟到表为:{}'.format(name1,times1,times2))

print()
for row in ws.iter_rows(min_row=2, values_only=True):
    for row2 in wws.iter_rows(min_row=3, values_only=True):
        name1 = row[1]
        name2 = row2[1]
        if name1 == name2:
            break
    if name1 != name2:
        print('迟到表中缺少'+name1+'的信息')
        ws2.append(row)
wb2.save('./迟到人员缺失.xlsx')

print()
for row in wws.iter_rows(min_row=3, values_only=True):
    times = row[13]
    name = row[1]
    if times == 0:
        print('{}为全勤!'.format(name))
