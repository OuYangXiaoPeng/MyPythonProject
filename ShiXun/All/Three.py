import win32com.client as win32
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border


# 第三步，为方便表格的样式修改
# 把xls表转xlsx表
fname = "E:\\pythonProject\\ShiXun\\All\\热门榜Top100.xls"#绝对路径
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wb.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()


# 由于依赖的包发生较大的版本升级，那么往往会出现兼容性问题，引起编译器警告或报错。
# 引入这段代码可以防止报错
import warnings
warnings.filterwarnings("ignore")


# 定义表头颜色样式为蓝色
header_fill = PatternFill('solid',fgColor='5B9BD5')
# 定义表中颜色样式为淡黄色
content_fill = PatternFill('solid',fgColor='FFFFE0')
# 定义表中颜色样式为黄色
content_fill2 = PatternFill('solid',fgColor='FFFF00')
# 定义对齐样式横向居中、纵向居中
align = Alignment(horizontal='center',vertical='center')
# 定义边样式为细条
side = Side('thin')
# 定义边样式，有底边和右边
header_border = Border(bottom=side,right=side)
# 定义表中边框样式，有左,右两边和底边
content_border = Border(bottom=side,right=side,left=side)
# 打开工作表
wb = load_workbook('热门榜Top100.xlsx')
# 获取活动表
ws = wb.active
# 获取最后一行行号
row_num = ws.max_row
# 循环第一行单元格，调整表头样式
for cell in ws[1]:
    cell.fill = header_fill
    cell.alignment = align
    cell.border = header_border
# 定义num方便换色
num = 0
# 从第二行开始循环修改样式
for row in ws.iter_rows(min_row=2,max_row=row_num):
    # 循环取出单元格，调整表中样式
    for cell in row:
        #如果num%5!=0则换一种颜色填充
        if num % 5 != 0:
            cell.fill = content_fill
        else:
            cell.fill = content_fill2
        cell.alignment = align
        cell.border = content_border
    num = num + 1
#  保存
wb.save('热门榜Top100.xlsx')

# 生成图表
wb = load_workbook('热门榜Top100.xlsx')
# 读取工作簿中的活跃工作表
ws = wb.active
# 实例化 LineChart类，得到LineChart对象
# Top10视频播放量的柱状图
chart = BarChart()
# 引用工作表的部分数据
data = Reference(worksheet=ws, min_row=2, max_row=11, min_col=5, max_col=5)
# 添加被引用的数据到LineChart对象
chart.add_data(data, from_rows=False, titles_from_data=False)
# 添加LineChart对象到工作表中，指定折线图的位置
ws.add_chart(chart, "G2")
# 引用工作表的表头数据
cats = Reference(worksheet=ws, min_row=2, max_row=11, min_col=2, max_col=2)
# 设置类别轴的标签
chart.set_categories(cats)
# 设置x轴的标题
chart.x_axis.title = "视频TOP10"
# 设置y轴的标题
chart.y_axis.title = "播放量(万)"
# Top10视频排行的折线图
chart = LineChart()
# 引用工作表的部分数据
data = Reference(worksheet=ws, min_row=2, max_row=11, min_col=1, max_col=1)
# 添加被引用的数据到LineChart对象
chart.add_data(data, from_rows=False, titles_from_data=False)
# 添加LineChart对象到工作表中，指定折线图的位置
ws.add_chart(chart, "G20")
# 引用工作表的表头数据
cats = Reference(worksheet=ws, min_row=2, max_row=11, min_col=2, max_col=2)
# 设置类别轴的标签
chart.set_categories(cats)
# 设置x轴的标题
chart.x_axis.title = "视频TOP10"
# 设置y轴的标题
chart.y_axis.title = "排名"
# 改变线条颜色
chart.style = 10
chart.style = 18
chart.style = 28
# 保存文件
wb.save('热门榜Top100.xlsx')