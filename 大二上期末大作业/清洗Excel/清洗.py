import xlrd as xd
from openpyxl.styles import PatternFill, Alignment, Side, Border

data =xd.open_workbook ('期末作业.xls') #打开excel表所在路径
sheet = data.sheet_by_index(0)  #读取数据，以excel表名来打开
d = []
for r in range(sheet.nrows): #将表中数据按行逐步添加到列表中，最后转换为list结构
    data1 = []
    for c in range(sheet.ncols):
        data1.append(sheet.cell_value(r,c))
    d.append(list(data1))
list=[]
hang=[]
for i in range(1,31):
    list.append(d[i][2].split('|'))
    hang.append(list[i-1][4].split('/'))

b=[]
for i in range(1,31):
    b=d[i][0]
    a=b[11:]
    d[i].append(a)



import xlwt
workbook = xlwt.Workbook(encoding='utf-8', style_compression=0)
worksheet = workbook.add_sheet('sheet1', cell_overwrite_ok=True)
# 加入表头
worksheet.write(0,0,"编号")
worksheet.write(0,1,"号码")
worksheet.write(0,2,"网址")
worksheet.write(0,3,"简介")
worksheet.write(0,4,"户型")
worksheet.write(0,5,"面积")
worksheet.write(0,6,"朝向")
worksheet.write(0,7,"建成时间及楼层数")
worksheet.write(0,8,"最近单价")
for i in range(1, 31):
    worksheet.write(i , 0, d[i][3])
    worksheet.write(i , 1, d[i][0])
    worksheet.write(i , 2, d[i][1])
    worksheet.write(i,3,list[i-1][0])
    worksheet.write(i,4,list[i-1][1])
    worksheet.write(i,5,list[i-1][2])
    worksheet.write(i,6,list[i-1][3])
    worksheet.write(i,7,hang[i-1][0])
    worksheet.write(i,8,hang[i-1][1])
# first_col = worksheet.col(2)  # 从零开始算，第一列
# first_col.width = 256 * 50
workbook.save('测试.xls')

import pandas as pd
from openpyxl import load_workbook
df = pd.read_excel("测试.xls", index_col='编号')
df.sort_values(by='编号',inplace=True,ascending=True)
df.to_excel('清洗完成品.xlsx')

wb = load_workbook('清洗完成品.xlsx')
sheet = wb.active
header_fill = PatternFill('solid',fgColor='5B9BD5')
# 定义表中颜色样式为淡黄色
content_fill = PatternFill('solid',fgColor='FFFFE0')
# 定义表中颜色样式为黄色
content_fill2 = PatternFill('solid',fgColor='FFFF00')
# 定义对齐样式横向居中、纵向居中
align = Alignment(horizontal='center',vertical='center')
# 定义边样式为细条
side = Side('thin')
# 定义边样式，有底边和右边
header_border = Border(bottom=side,right=side)
# 定义表中边框样式，有左,右两边和底边
content_border = Border(bottom=side,right=side,left=side)
# 获取最后一行行号
row_num = sheet.max_row
# 循环第一行单元格，调整表头样式
for cell in sheet[1]:
    cell.fill = header_fill
    cell.alignment = align
    cell.border = header_border
# 定义num方便换色
num = 0
# 从第二行开始循环修改样式
for row in sheet.iter_rows(min_row=2,max_row=row_num):
    # 循环取出单元格，调整表中样式
    for cell in row:
        #如果num%5!=0则换一种颜色填充
        if num % 5 != 0:
            cell.fill = content_fill
        else:
            cell.fill = content_fill2
        cell.alignment = align
        cell.border = content_border
    num = num + 1

sheet.column_dimensions["B"].width = 25
sheet.column_dimensions["C"].width = 50
sheet.column_dimensions["D"].width = 50
sheet.column_dimensions["F"].width = 15
sheet.column_dimensions["H"].width = 50
sheet.column_dimensions["I"].width = 180
wb.save('清洗完成品.xlsx')
