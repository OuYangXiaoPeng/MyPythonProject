import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border

head = PatternFill('solid', fgColor='C5D9F1')  # 定义表头颜色样式为橙色
head1 = PatternFill('solid', fgColor='FF7F24')  # 定义表头颜色样式为橙色
head2 = PatternFill('solid', fgColor='FFFF00')  # 定义表头颜色样式为橙色
content = PatternFill('solid', fgColor='FFFFE0')  # 定义表中、表尾颜色样式为淡黄色
bottom = PatternFill('solid', fgColor='EE9572')  # 定义表尾颜色样式为淡桔红色
align = Alignment(horizontal='center', vertical='center')  # 定义对齐样式横向居中、纵向居中
side = Side('thin')  # 定义边样式为细条
border = Border(bottom=side, right=side)  # 定义表头边框样式，有底边和右边
content1 = Border(left=side)  # 定义表中、表尾边框样式，有左边
path = "./21"  # 设置文件夹路径
file = os.listdir(path)
for i in file:
    j = path + "/" + i
    wb = load_workbook(j)
    ws = wb.active
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    for m in ws[1]:
        m.fill = head1
        m.alignment = align
        m.border = border
    row_num = ws.max_row

    num = 0;
    for i in ws.iter_rows(min_row=2, max_row=(row_num)):
        if (i[9].value == "欧阳小鹏"):
            break
        num += 1

    for n in ws.iter_cols(min_row=2, max_row=(row_num)):
        for p in n:
            p.fill = head
            p.alignment = align
            p.border = border
        n[num].fill = head2
    wb.save(j)
