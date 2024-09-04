import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Alignment,Side,Border

# 定义表头颜色样式为橙色
header_fill = PatternFill('solid',fgColor='FF9900')
# 定义表中、表尾颜色样式为淡黄色
content_fill = PatternFill('solid',fgColor='FFFFE0')
# 定义表自己行背景色
content_fill2 = PatternFill('solid',fgColor='BDD7EE')

# 定义对齐样式横向居中、纵向居中
align = Alignment(horizontal='center',vertical='center')

# 定义边样式为细条
side = Side('thin')

# 定义边样式，有底边和右边
header_border = Border(bottom=side,right=side)
# 定义表中、表尾边框样式，有左边
content_border = Border(left=side)

# 设置文件夹路径
path='./21计应301.xlsx'
# 循环文件名列表
wb = load_workbook(path)
# 打开工作表
ws = wb.active

# 循环第一行单元格，调整表头样式
for cell in ws[1]:
    # 设置单元格填充颜色
    cell.fill = header_fill
    # 设置单元格对齐方式
    cell.alignment = align
    # 设置单元格边框
    cell.border = header_border

# 获取最后一行行号
row_num = ws.max_row

num = 0;
for i in ws.iter_rows(min_row=2, max_row=(row_num)):
    if (i[9].value == "欧阳小鹏"):
        break
    num += 1
# 从第二行开始，循环倒数第二行
for col in ws.iter_cols(min_row=2,max_row=row_num):
    # 循环取出单元格，调整表中样式
    for cell in col:
        cell.fill = content_fill
        cell.alignment = align
        cell.border = content_border
        cell.border = header_border
    col[num].fill = content_fill2

# num = 0
# for row in ws['j']:
#     num += 1
#     if (row.value =="欧阳小鹏"):
#             for cell in ws[num]:
#                 #设置单元格填充颜色
#                 cell.fill = content_fill2
#                 cell.alignment = align#设置单元格左边框
#                 cell.border = header_border#设置单元格低边框
#     else:
#         continue


# 保存
wb.save(path)

