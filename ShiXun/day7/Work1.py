from openpyxl import load_workbook, Workbook
# 打开【10月考勤统计.xlsx】工作簿
wb = load_workbook('./21计应301.xlsx')
#获取活动工作表
ws = wb.active
# 获取表头
late_header = []
for cell in ws[1]:
    late_header.append(cell.value)
# 新建工作簿
    new_wb1 = Workbook()
    new_wb2 = Workbook()
    new_wb3 = Workbook()
    new_wb4 = Workbook()
    new_wb5 = Workbook()
#获取新工作簿中的工作表
new_ws1 = new_wb1.active
new_ws2 = new_wb2.active
new_ws3 = new_wb3.active
new_ws4 = new_wb4.active
new_ws5 = new_wb5.active
#将表头写入新工作簿的工作表中
new_ws1.append(late_header)
new_ws2.append(late_header)
new_ws3.append(late_header)
new_ws4.append(late_header)
new_ws5.append(late_header)
# 从第二行开始遍历表格
for row in ws.iter_rows(min_row=2, values_only=True):
    #取出姓名，迟到时间和迟到次数
    name = row[-3]
    score = row[-1]
    # 录入成绩
    if score >= 90:
        print('{}成绩为:{}'.format(name, score))
        new_ws1.append(row)
    elif score >= 80 and score < 90:
        print('{}成绩为:{}'.format(name, score))
        new_ws2.append(row)
    elif score >= 70 and score < 80:
        print('{}成绩为:{}'.format(name, score))
        new_ws3.append(row)
    elif score >= 60 and score < 70:
        print('{}成绩为:{}'.format(name, score))
        new_ws4.append(row)
    else:
        print('{}成绩为:{}'.format(name, score))
        new_ws5.append(row)
# 将新工作簿保存为【成绩表.xlsx】
new_wb1.save('./学生成绩等级/优秀学生信息表.xlsx')
new_wb2.save('./学生成绩等级/良好学生成绩表.xlsx')
new_wb3.save('./学生成绩等级/中等学生成绩表.xlsx')
new_wb4.save('./学生成绩等级/及格学生成绩表.xlsx')
new_wb5.save('./学生成绩等级/需补考学生成绩表.xlsx')

