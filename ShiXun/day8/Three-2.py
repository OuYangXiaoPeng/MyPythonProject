from openpyxl import load_workbook, Workbook
import re
import string

import requests
from openpyxl.styles import PatternFill, Alignment, Border, Side

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'}
url = 'http://front-gateway.mtime.com/community/top_list/detail.api?tt=1668579076556&id=300773&pageIndex=1&pageSize=10'
res = requests.get(url, headers=headers, timeout=20)
res.encoding = 'utf-8'
HtmlStr = res.text
print(HtmlStr)
PianMing = re.findall('"movieName":(.*?),"movieNameEn"', HtmlStr)
DaoYan = re.findall('"director":(.*?),"actors"', HtmlStr)
shijian = re.findall('"releaseDate":(.*?),"releaseLocation"', HtmlStr)
chandi = re.findall('"releaseLocation":(.*?),"score"', HtmlStr)

out=[]
for i in PianMing:
    tmp = str.maketrans({key: None for key in string.punctuation})
    j = i.translate(tmp)
    out.append(j)
print(out)
out1=[]
for i in DaoYan:
    tmp = str.maketrans({key: None for key in string.punctuation})
    j = i.translate(tmp)
    out1.append(j)
print(out1)
out2=[]
for i in shijian:
    tmp = str.maketrans({key: None for key in string.punctuation})
    j = i.translate(tmp)
    out2.append(j)
print(out2)
out3=[]
for i in chandi:
    tmp = str.maketrans({key: None for key in string.punctuation})
    j = i.translate(tmp)
    out3.append(j)
print(out3)

import xlwt
late_header = ["排名","电影名","导演","时间","产地"]
f = xlwt.Workbook('encoding = utf-8') #设置工作簿编码
sheet1 = f.add_sheet('sheet1',cell_overwrite_ok=True) #创建sheet工作表
for i in range(len(late_header)):
    sheet1.write(0,i,late_header[i])
for i in range(len(out)):
    sheet1.write(i+1,0,i+1) #写入数据参数对应 行, 列, 值
    sheet1.write(i+1,1,out[i]) #写入数据参数对应 行, 列, 值
    sheet1.write(i+1,2,out1[i]) #写入数据参数对应 行, 列, 值
    sheet1.write(i+1,3,out2[i]) #写入数据参数对应 行, 列, 值
    sheet1.write(i+1,4,out3[i]) #写入数据参数对应 行, 列, 值
sec_col=sheet1.col(1)
sec_col.width=256*25
fuor_col=sheet1.col(3)
fuor_col.width=256*25
f.save('./信息.xls')

import win32com.client as win32
fname = "E:\\pythonProject\\ShiXun\\day8\\信息.xls"#绝对路径
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(fname)
wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wb.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()

# 一旦依赖的包发生较大的版本升级，那么往往会出现兼容性问题， 引起编译器警告或报错。
import warnings
warnings.filterwarnings("ignore")

header_fill = PatternFill('solid',fgColor='5B9BD5')
content_fill = PatternFill('solid',fgColor='FFFFE0')
align = Alignment(horizontal='center',vertical='center')
side = Side('thin')
header_border = Border(bottom=side,right=side)
content_border = Border(left=side)
# 打开工作表
wb = load_workbook('信息.xlsx')
ws = wb.active
row_num = ws.max_row
for cell in ws[1]:
    cell.fill = header_fill
    cell.alignment = align
    cell.border = header_border
for row in ws.iter_rows(min_row=2,max_row=row_num):
    for cell in row:
        cell.fill = content_fill
        cell.alignment = align
        cell.border = content_border
wb.save('信息.xlsx')

